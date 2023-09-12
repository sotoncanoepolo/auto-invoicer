from dataclasses import dataclass

import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet

from core import Person, InvoiceItem


@dataclass
class Header:
    name: str
    column: int


class Sheet:
    def __init__(self, spreadsheet_path: str, sheet_name: str = 'Main'):
        wb: xl.Workbook = xl.load_workbook(filename=spreadsheet_path, data_only=True)
        self.ws: Worksheet = wb[sheet_name]

        # Constants
        self.header_row = 2
        self.name_header = self.expect_header("Name")
        self.refund_header = self.expect_header("Refunds")
        self.total_cost_header = self.expect_header("Total Cost")
        self.paid_header = self.expect_header("Paid?")
        self.notes_header = self.expect_header("Notes")
        normal_headers = ["Deposits", "Entry", "Camping", "Food", "Driver Debt", "Cost towards travel", "Travel Costs",
                          "Ferry Costs"]
        self.normal_headers: list[Header] = list(
            filter(lambda h: h is not None, map(lambda n: self.get_header(n), normal_headers)))

    def get_header(self, name: str) -> Header | None:
        for header in self.ws[self.header_row]:
            if header.value == name:
                return Header(name, header.column)
        print(f'Did not find header with name {name}, ignore if the header should not exist.')
        return None

    def expect_header(self, name: str) -> Header:
        header = self.get_header(name)
        if header is None:
            raise Exception(f'Could not find header with name {name} on row {self.header_row}.')
        return header

    def get_people(self) -> list[Person]:
        people = []
        for row in range(self.header_row + 1, self.ws.max_row):
            cell = self.ws.cell(row, self.name_header.column)
            if not isinstance(cell.value, str):
                break
            if cell.value.strip() == "":
                continue
            if self._has_paid(row, self.paid_header.column):
                continue
            if self._get_expense(row, self.total_cost_header) <= 0:
                continue

            people.append(Person(cell.value, row))

        return people

    def _has_paid(self, row: int, col: int) -> bool:
        cell = self.ws.cell(row, col)
        if not isinstance(cell.value, bool):
            raise Exception(f'Expected a bool at ({row}, {col}) indicating whether the person has paid.')
        return cell.value

    def _get_expense(self, row: int, header: Header) -> int:
        cell = self.ws.cell(row, header.column)
        if cell.value is None:
            return 0
        if not (isinstance(cell.value, float) or isinstance(cell.value, int)):
            raise Exception(
                f'Expected a number at ({row}, {header.column}) indicating the cost for header: {header.name}.')
        return int(round(cell.value * 100))

    def add_costs(self, person: Person):
        row = person.row

        person.costs = list(filter(lambda i: i.cost != 0,
                                   map(lambda h: InvoiceItem(h.name, h.name, self._get_expense(row, h)),
                                       self.normal_headers)))

        refund_cost = self._get_expense(row, self.refund_header)

        if refund_cost != 0:
            refund_description = self.ws.cell(row, self.notes_header.column).value
            if refund_description is None or str(refund_description).strip() == "":
                refund_description = self.refund_header.name

            refund = InvoiceItem(self.refund_header.name, str(refund_description), refund_cost)
            person.costs.append(refund)

        total_cost = sum(map(lambda i: i.cost, person.costs))
        real_total_cost = self._get_expense(row, self.total_cost_header)
        person.total_cost = real_total_cost
        if total_cost != real_total_cost:
            diff = real_total_cost - total_cost
            round_error = InvoiceItem("Rounding Error", "The total does not match the total of " +
                                      "adding all items without rounding so this " +
                                      "is added to fix the rounding error.", diff)
            person.costs.append(round_error)
